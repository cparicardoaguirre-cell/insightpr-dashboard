#!/usr/bin/env python3
"""
Financial Statement Data Extractor for NLT-PR Dashboard
========================================================
This script extracts financial data from the Excel workbook and generates
structured JSON files for the dashboard, organizing leadschedules by category.

Author: Auto-generated for NLT-PR Dashboard
Version: 1.0.0
"""

import json
import openpyxl
import warnings
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional

# Suppress openpyxl warnings
warnings.filterwarnings('ignore')

# Configuration
EXCEL_FILE = Path("D:/NLTS-PR/NLTS-PR FS 12 31 2024 Rev156-3.xlsm")
OUTPUT_DIR = Path("c:/Users/cpari/.gemini/antigravity/NLT_PR_Dashboard/src/data")

# Sheet mappings for financial statements
SHEET_CONFIG = {
    "BS": {"sheet": "BS", "type": "standard", "name_en": "Balance Sheet", "name_es": "Estado de Situación"},
    "IS": {"sheet": "IS", "type": "standard", "name_en": "Income Statement", "name_es": "Estado de Resultados"},
    "CF": {"sheet": "CF", "type": "standard", "name_en": "Cash Flow", "name_es": "Flujo de Efectivo"},
    "Lead": {"sheet": "Leadschedules", "type": "grid", "name_en": "Leadschedules", "name_es": "Cédulas Contables"},
    "TaxLead": {"sheet": "TaxLeadschedules", "type": "grid", "name_en": "Tax Leadschedules", "name_es": "Cédulas Fiscales"},
}

# Leadschedule section keywords for categorization
LEADSCHEDULE_SECTIONS = {
    "assets": {
        "name_en": "Assets",
        "name_es": "Activos",
        "icon": "📦",
        "keywords": ["Cash", "Receivable", "Inventory", "Prepaid", "Property", "Equipment", 
                    "Machinery", "Vehicles", "Depreciation", "Fixed Assets", "Current Assets",
                    "Efectivo", "Cuentas por Cobrar", "Inventario", "Activos"]
    },
    "liabilities": {
        "name_en": "Liabilities", 
        "name_es": "Pasivos",
        "icon": "📋",
        "keywords": ["Payable", "Accrued", "Loan", "Debt", "Lease", "Credit", "Due to",
                    "Cuentas por Pagar", "Préstamo", "Deuda", "Pasivos"]
    },
    "equity": {
        "name_en": "Shareholders' Equity",
        "name_es": "Capital Contable",
        "icon": "🏛️",
        "keywords": ["Equity", "Capital", "Stock", "Retained Earnings", "Dividends",
                    "Acciones", "Capital", "Utilidades Retenidas"]
    },
    "revenue": {
        "name_en": "Revenue",
        "name_es": "Ingresos",
        "icon": "💰",
        "keywords": ["Revenue", "Sales", "Income", "Service", "Rental",
                    "Ingresos", "Ventas", "Servicios", "Alquiler"]
    },
    "costs": {
        "name_en": "Costs of Revenue",
        "name_es": "Costo de Ventas",
        "icon": "🔧",
        "keywords": ["Cost of", "COGS", "Material", "Labor", "Direct",
                    "Costo de", "Material", "Mano de Obra"]
    },
    "expenses": {
        "name_en": "Operating Expenses",
        "name_es": "Gastos Operacionales", 
        "icon": "📊",
        "keywords": ["Expense", "Rent", "Utilities", "Insurance", "Professional",
                    "Advertising", "Travel", "Office", "Bank Charges", "Computer",
                    "Gasto", "Alquiler", "Seguros", "Servicios Profesionales"]
    },
    "other": {
        "name_en": "Other Income/Expenses",
        "name_es": "Otros Ingresos/Gastos",
        "icon": "📈",
        "keywords": ["Interest", "Other Income", "Other Expense", "Gain", "Loss",
                    "Intereses", "Otros Ingresos", "Otros Gastos", "Ganancia", "Pérdida"]
    },
    "taxes": {
        "name_en": "Taxes & Contributions",
        "name_es": "Contribuciones",
        "icon": "🏦",
        "keywords": ["Tax", "Contribución", "IVU", "Patente", "Municipal", "Arbitrios",
                    "Alternativa Mínima", "Income Tax", "Provision"]
    }
}


def get_cell_value(cell) -> Any:
    """Extract value from cell, handling various types."""
    if cell is None:
        return ""
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def get_cell_format(cell) -> str:
    """Extract number format from cell."""
    if cell is None:
        return "General"
    try:
        fmt = cell.number_format
        return fmt if fmt else "General"
    except:
        return "General"


def extract_grid_data(ws, max_rows: int = 500, max_cols: int = 15) -> List[List[Dict]]:
    """Extract grid data from a worksheet."""
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols), 1):
        row_data = []
        has_value = False
        for cell in row:
            value = get_cell_value(cell)
            if value:
                has_value = True
            row_data.append({
                "v": value,
                "f": get_cell_format(cell)
            })
        if has_value:
            rows.append(row_data)
    return rows


def categorize_row(row: List[Dict], sections: Dict) -> str:
    """Categorize a row based on its content."""
    first_cell = str(row[0].get("v", "")).lower() if row else ""
    
    for section_key, section_info in sections.items():
        for keyword in section_info["keywords"]:
            if keyword.lower() in first_cell:
                return section_key
    
    return "other"


def extract_leadschedules_by_section(ws, sections: Dict) -> Dict[str, List[List[Dict]]]:
    """Extract leadschedule data organized by section."""
    all_rows = extract_grid_data(ws)
    
    categorized = {key: [] for key in sections.keys()}
    header_rows = []
    
    # First few rows are typically headers
    for i, row in enumerate(all_rows[:5]):
        header_rows.append(row)
    
    # Categorize remaining rows
    current_section = None
    for row in all_rows[5:]:
        first_cell = str(row[0].get("v", "")) if row else ""
        
        # Check if this is a section header
        section = categorize_row(row, sections)
        
        # If row has content, add to appropriate section
        if any(cell.get("v", "") for cell in row):
            categorized[section].append(row)
            current_section = section
    
    # Add headers to each section that has data
    for section_key, rows in categorized.items():
        if rows:
            categorized[section_key] = header_rows + rows
    
    return categorized


def extract_standard_statement(ws, start_row: int = 4, end_row: int = 100) -> List[Dict]:
    """Extract standard financial statement data (BS, IS, CF)."""
    items = []
    current_section = "General"
    
    for row_idx in range(start_row, end_row + 1):
        name = get_cell_value(ws.cell(row=row_idx, column=1))
        if not name:
            continue
            
        # Detect section headers (typically bold or indented differently)
        col2 = get_cell_value(ws.cell(row=row_idx, column=2))
        col3 = get_cell_value(ws.cell(row=row_idx, column=3))
        
        # Check if this is a section header
        name_str = str(name).strip()
        if name_str.isupper() or (not col2 and not col3 and name_str):
            current_section = name_str
            
        item = {
            "name": name,
            "2024": get_cell_value(ws.cell(row=row_idx, column=2)) or 0,
            "2023": get_cell_value(ws.cell(row=row_idx, column=3)) or 0,
            "section": current_section,
            "row_hidden": ws.row_dimensions[row_idx].hidden if hasattr(ws.row_dimensions[row_idx], 'hidden') else False,
            "format": get_cell_format(ws.cell(row=row_idx, column=2)),
            "indent": 0
        }
        items.append(item)
    
    return items


def main():
    """Main extraction function."""
    print(f"📊 Loading workbook: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    output_data = {
        "Metadata": {
            "SourceFile": EXCEL_FILE.name,
            "ExtractedAt": datetime.now().isoformat(),
            "PdfAvailable": True,
            "Version": "2.0.0"
        },
        "Sections": {}
    }
    
    # Extract Leadschedules by section
    print("📋 Extracting Leadschedules...")
    if "Leadschedules" in wb.sheetnames:
        ws = wb["Leadschedules"]
        lead_sections = extract_leadschedules_by_section(ws, LEADSCHEDULE_SECTIONS)
        output_data["LeadSections"] = {}
        for section_key, rows in lead_sections.items():
            if rows:
                section_info = LEADSCHEDULE_SECTIONS[section_key]
                output_data["LeadSections"][section_key] = {
                    "name_en": section_info["name_en"],
                    "name_es": section_info["name_es"],
                    "icon": section_info["icon"],
                    "data": rows
                }
                print(f"  ✓ {section_info['name_en']}: {len(rows)} rows")
    
    # Extract Tax Leadschedules by section    
    print("🏦 Extracting Tax Leadschedules...")
    if "TaxLeadschedules" in wb.sheetnames:
        ws = wb["TaxLeadschedules"]
        tax_lead_sections = extract_leadschedules_by_section(ws, LEADSCHEDULE_SECTIONS)
        output_data["TaxLeadSections"] = {}
        for section_key, rows in tax_lead_sections.items():
            if rows:
                section_info = LEADSCHEDULE_SECTIONS[section_key]
                output_data["TaxLeadSections"][section_key] = {
                    "name_en": section_info["name_en"],
                    "name_es": section_info["name_es"],
                    "icon": section_info["icon"],
                    "data": rows
                }
                print(f"  ✓ {section_info['name_en']}: {len(rows)} rows")
    
    # Also keep the full grid for backward compatibility
    print("📄 Extracting full grids...")
    if "Leadschedules" in wb.sheetnames:
        output_data["Lead"] = extract_grid_data(wb["Leadschedules"])
        print(f"  ✓ Lead: {len(output_data['Lead'])} rows")
        
    if "TaxLeadschedules" in wb.sheetnames:
        output_data["TaxLead"] = extract_grid_data(wb["TaxLeadschedules"])
        print(f"  ✓ TaxLead: {len(output_data['TaxLead'])} rows")
    
    # Keep existing BS, IS, CF data if present in current file
    print("📊 Preserving standard statements from existing file...")
    existing_file = OUTPUT_DIR / "financial_statements.json"
    if existing_file.exists():
        with open(existing_file, 'r', encoding='utf-8') as f:
            existing = json.load(f)
            for key in ["BS", "IS", "CF"]:
                if key in existing:
                    output_data[key] = existing[key]
                    print(f"  ✓ {key}: preserved")
    
    # Save output
    output_file = OUTPUT_DIR / "financial_statements.json"
    print(f"\n💾 Saving to: {output_file}")
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)
    
    print("✅ Extraction complete!")
    
    # Print summary
    print("\n📊 Summary:")
    if "LeadSections" in output_data:
        print(f"  - Lead Sections: {len(output_data['LeadSections'])}")
    if "TaxLeadSections" in output_data:
        print(f"  - Tax Lead Sections: {len(output_data['TaxLeadSections'])}")


if __name__ == "__main__":
    main()
