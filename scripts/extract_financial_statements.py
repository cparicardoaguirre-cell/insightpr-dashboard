import openpyxl
import json
import os
import shutil
import glob
import re
import warnings

# Suppress warnings
warnings.filterwarnings("ignore", category=UserWarning)

SOURCE_DIR = r"D:\NLTS-PR"
OUTPUT_JSON = r"src/data/financial_statements.json"
PUBLIC_DOCS_DIR = r"public/documents"

def find_latest_files():
    # Find all Excel files matching pattern
    excel_pattern = os.path.join(SOURCE_DIR, "NLTS-PR FS*.xlsm")
    excel_files = glob.glob(excel_pattern)
    
    if not excel_files:
        print("No Excel files found.")
        return None, None

    # Sort by revision number (e.g., Rev156-3)
    def parse_revision(fname):
        match = re.search(r"Rev(\d+)-(\d+)", fname)
        if match:
            return int(match.group(1)) * 1000 + int(match.group(2))
        return 0

    latest_excel = max(excel_files, key=parse_revision)
    print(f"Latest Excel found: {latest_excel}")

    # Find matching PDF (Look in subfolders too)
    # Strategy: Look for strict name match first, then same revision
    base_name = os.path.splitext(os.path.basename(latest_excel))[0]
    
    # 1. Look for exact PDF name in ROOT
    pdf_candidates = glob.glob(os.path.join(SOURCE_DIR, f"{base_name}.pdf"))
    
    # 2. Look for exact PDF name in 2024 subfolder
    if not pdf_candidates:
        pdf_candidates = glob.glob(os.path.join(SOURCE_DIR, "**", f"{base_name}.pdf"), recursive=True)
    
    latest_pdf = pdf_candidates[0] if pdf_candidates else None
    
    if latest_pdf:
        print(f"Matching PDF found: {latest_pdf}")
    else:
        print("No matching PDF found.")

    return latest_excel, latest_pdf

def extract_financials():
    latest_excel, latest_pdf = find_latest_files()
    
    if not latest_excel:
        return

    # Create public docs dir
    if not os.path.exists(PUBLIC_DOCS_DIR):
        os.makedirs(PUBLIC_DOCS_DIR)

    # Copy files to public
    print(f"Copying {latest_excel} to public...")
    shutil.copy2(latest_excel, os.path.join(PUBLIC_DOCS_DIR, "latest_source.xlsm"))
    if latest_pdf:
        print(f"Copying {latest_pdf} to public...")
        shutil.copy2(latest_pdf, os.path.join(PUBLIC_DOCS_DIR, "audited_financials.pdf"))

    print(f"Loading workbook: {latest_excel}")
    try:
        wb = openpyxl.load_workbook(latest_excel, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    financials = {
        "BS": [],
        "IS": [],
        "CF": [],
        "TaxLead": [],
        "Lead": [],
        "Metadata": {
            "SourceFile": os.path.basename(latest_excel),
            "PdfAvailable": bool(latest_pdf)
        }
    }

    # Helper function for standard 5-col extraction (BS/IS/CF)
    def extract_standard_sheet(sheet_name, target_key, col_map):
        if sheet_name not in wb.sheetnames:
            print(f"Skipping {sheet_name} (Not found)")
            return

        print(f"Processing {sheet_name}...")
        ws = wb[sheet_name]
        current_section = "General"
        
        for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True)):
            if not row or len(row) < 5: continue
            
            desc = row[col_map['desc']]
            val_2024 = row[col_map['2024']]
            val_2023 = row[col_map['2023']]
            
            # Debug CF
            if sheet_name == "CF":
                if desc and (val_2024 != None or val_2023 != None):
                     pass
                     # print(f"CF Row {i+5}: Desc='{desc}', 24={val_2024}, 23={val_2023}")

            if not desc: continue
            
            desc_str = str(desc).strip()
            
            # Heuristic Section detection
            if desc_str.isupper() and len(desc_str) > 4:
                current_section = desc_str
                # Usually headers
                if val_2024 is None and val_2023 is None:
                    continue

            # Skip empty value rows
            if val_2024 is None and val_2023 is None:
                continue

            def clean_val(v):
                if isinstance(v, (int, float)): return round(v)
                return 0

            item = {
                "name": desc_str,
                "2024": clean_val(val_2024),
                "2023": clean_val(val_2023),
                "section": current_section
            }

            if item["2024"] != 0 or item["2023"] != 0:
                financials[target_key].append(item)

    # 1. BS Extraction (Desc:0, 24:2, 23:4)
    extract_standard_sheet("BS", "BS", {'desc': 0, '2024': 2, '2023': 4})
    
    # 2. IS Extraction (Desc:0, 24:2, 23:4)
    extract_standard_sheet("IS", "IS", {'desc': 0, '2024': 2, '2023': 4})

    # 3. CF Extraction (Desc:0, 24:2, 23:4)
    extract_standard_sheet("CF", "CF", {'desc': 0, '2024': 2, '2023': 4})
    
    # 4. Tax Leadschedule / Leadschedule
    def extract_table_sheet(sheet_keyword, target_key):
        sheet_name = next((s for s in wb.sheetnames if sheet_keyword.lower() == s.lower()), None)
        if not sheet_name:
            print(f"Skipping {sheet_keyword} (Not found)")
            return

        print(f"Processing {sheet_name} into {target_key}...")
        ws = wb[sheet_name]
        
        table_data = []
        for row in ws.iter_rows(min_row=1, max_row=200, max_col=12, values_only=True):
            clean_row = [str(c).strip() if c is not None else "" for c in row]
            if any(clean_row):
                table_data.append(clean_row)
        
        financials[target_key] = table_data

    extract_table_sheet("TaxLeadschedules", "TaxLead")
    extract_table_sheet("Leadschedules", "Lead")

    # Save
    if not os.path.exists(os.path.dirname(OUTPUT_JSON)):
        os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
        
    with open(OUTPUT_JSON, "w") as f:
        json.dump(financials, f, indent=2)

    print(f"Extraction complete. Saved to {OUTPUT_JSON}")
    print(f"BS Items: {len(financials['BS'])}")
    print(f"IS Items: {len(financials['IS'])}")
    print(f"CF Items: {len(financials['CF'])}")

if __name__ == "__main__":
    extract_financials()
