import openpyxl

SOURCE_FILE = r"D:\NLTS-PR\NLTS-PR FS 12 31 2024 Rev156-3.xlsm"

def inspect_structure():
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    
    print(f"Sheet Names: {wb.sheetnames}")
    
    target_sheets = ["CF", "Tax Leadschedule", "Leadschedule", "Tax", "Lead"] # Potential names
    
    for sheet_name in wb.sheetnames:
        # Fuzzy match or exact match check
        if any(target in sheet_name for target in ["CF", "Tax", "Lead", "BS", "IS"]):
            print(f"\n--- {sheet_name} First 10 Rows ---")
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True, max_row=10))
            for i, row in enumerate(rows):
                print(f"Row {i+1}: {row[:6]}")

if __name__ == "__main__":
    inspect_structure()
