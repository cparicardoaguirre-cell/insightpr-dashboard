"""
Dynamic Financial Ratios Extractor for NLT-PR Dashboard
Automatically finds the latest NLTS-PR FS Excel file and extracts ratios.
"""
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

import json
import os
import re
from datetime import datetime
import openpyxl

NLTS_PR_DIR = r'D:\NLTS-PR'
OUTPUT_FILE = os.path.join(NLTS_PR_DIR, 'dynamic_ratios.json')

# Industry benchmarks for equipment rental/forklift service industry
# Sources: ReadyRatios (2023), United Rentals, Equipment Rental Industry Data
BENCHMARKS = {
    # Solvency Ratios
    'current ratio': {'good': 2.0, 'warning': 1.5, 'higher_is_better': True, 'industry': 2.31, 'industry_label': 'Equipment Rental Median 2023'},
    'quick ratio': {'good': 1.0, 'warning': 0.7, 'higher_is_better': True, 'industry': 1.2, 'industry_label': 'Industry Target'},
    
    # Safety Ratios
    'debt to equity': {'good': 2.0, 'warning': 3.0, 'higher_is_better': False, 'industry': 1.83, 'industry_label': 'ReadyRatios 2023'},
    'debt ratio': {'good': 0.50, 'warning': 0.70, 'higher_is_better': False, 'industry': 0.65, 'industry_label': 'Equipment Rental Median'},
    
    # Profitability Ratios
    'gross profit margin': {'good': 0.35, 'warning': 0.20, 'higher_is_better': True, 'industry': 0.40, 'industry_label': 'United Rentals 2024'},
    'operating margin': {'good': 0.15, 'warning': 0.08, 'higher_is_better': True, 'industry': 0.25, 'industry_label': 'United Rentals 2024'},
    'net profit margin before tax': {'good': 0.10, 'warning': 0.05, 'higher_is_better': True, 'industry': 0.08, 'industry_label': 'Equipment Rental 2024'},
    'net profit margin after tax': {'good': 0.08, 'warning': 0.03, 'higher_is_better': True, 'industry': 0.042, 'industry_label': 'Industry Median'},
    
    # Asset Management Ratios
    'sales to assets': {'good': 2.0, 'warning': 1.5, 'higher_is_better': True, 'industry': 1.5, 'industry_label': 'Industry Average'},
    'return on assets': {'good': 0.08, 'warning': 0.04, 'higher_is_better': True, 'industry': 0.065, 'industry_label': 'Industry Average'},
    'return on equity': {'good': 0.15, 'warning': 0.08, 'higher_is_better': True, 'industry': 0.12, 'industry_label': 'Industry Average'},
    'inventory turnover': {'good': 6.0, 'warning': 3.0, 'higher_is_better': True, 'industry': 5.0, 'industry_label': 'Equipment Rental Standard'},
    'inventory turnover (x)': {'good': 6.0, 'warning': 3.0, 'higher_is_better': True, 'industry': 5.0, 'industry_label': 'Equipment Rental Standard'},
    'inventory turnover (days)': {'good': 45, 'warning': 75, 'higher_is_better': False, 'industry': 73, 'industry_label': 'Industry Average'},
    'accounts receivable turnover': {'good': 10.0, 'warning': 6.0, 'higher_is_better': True, 'industry': 8.0, 'industry_label': 'Industry Average'},
    'accounts receivable turnover (x)': {'good': 10.0, 'warning': 6.0, 'higher_is_better': True, 'industry': 8.0, 'industry_label': 'Industry Average'},
    'collection period': {'good': 35, 'warning': 50, 'higher_is_better': False, 'industry': 45, 'industry_label': 'Industry Average'},
    'collection period (days)': {'good': 35, 'warning': 50, 'higher_is_better': False, 'industry': 45, 'industry_label': 'Industry Average'},
    'accounts payable turnover': {'good': 8.0, 'warning': 5.0, 'higher_is_better': True, 'industry': 7.0, 'industry_label': 'Industry Average'},
    'accounts payable turnover (x)': {'good': 8.0, 'warning': 5.0, 'higher_is_better': True, 'industry': 7.0, 'industry_label': 'Industry Average'},
    'accounts payable (days)': {'good': 50, 'warning': 70, 'higher_is_better': False, 'industry': 52, 'industry_label': 'Industry Average'},
    'personnel productivity': {'good': 0.50, 'warning': 0.70, 'higher_is_better': False, 'industry': 0.55, 'industry_label': 'Industry Target'},
    'gross margin return on inventory': {'good': 3.0, 'warning': 1.5, 'higher_is_better': True, 'industry': 2.5, 'industry_label': 'Retail/Distribution'},
    'interest coverage': {'good': 4.0, 'warning': 2.0, 'higher_is_better': True, 'industry': 4.5, 'industry_label': 'Equipment Rental 2024'},
    
    # Leading Financial Indicators
    'variable cost % of sales': {'good': 0.45, 'warning': 0.60, 'higher_is_better': False, 'industry': 0.50, 'industry_label': 'Industry Target'},
    'contribution margin': {'good': 0.50, 'warning': 0.35, 'higher_is_better': True, 'industry': 0.45, 'industry_label': 'Industry Standard'},
    'break-even as % of net sales': {'good': 0.70, 'warning': 0.85, 'higher_is_better': False, 'industry': 0.75, 'industry_label': 'Industry Target'},
    'break-even sales': {'good': None, 'warning': None, 'higher_is_better': False, 'industry': None, 'industry_label': 'Company-Specific'},
    '% break-even sales': {'good': 0.70, 'warning': 0.85, 'higher_is_better': False, 'industry': 0.75, 'industry_label': 'Industry Target'},
    'sustainable growth - current d/e': {'good': 0.15, 'warning': 0.05, 'higher_is_better': True, 'industry': 0.10, 'industry_label': 'Industry Average'},
    'sustainable growth - no new debt': {'good': 0.10, 'warning': 0.03, 'higher_is_better': True, 'industry': 0.06, 'industry_label': 'Industry Average'},
    'z-score': {'good': 3.0, 'warning': 1.8, 'higher_is_better': True, 'industry': 2.5, 'industry_label': 'Altman Safe Zone >2.99'},
    'z-score (bankruptcy indicator)': {'good': 3.0, 'warning': 1.8, 'higher_is_better': True, 'industry': 2.5, 'industry_label': 'Altman Safe Zone >2.99'},
    'retained earnings to total assets': {'good': 0.20, 'warning': 0.10, 'higher_is_better': True, 'industry': 0.15, 'industry_label': 'Industry Average'},
    'working capital to total assets': {'good': 0.15, 'warning': 0.05, 'higher_is_better': True, 'industry': 0.10, 'industry_label': 'Industry Average'},
    'ebitda': {'good': None, 'warning': None, 'higher_is_better': True, 'industry': None, 'industry_label': 'Company-Specific'},
}


def find_latest_fs_file():
    """
    Find the most recent NLTS-PR FS Excel file.
    Selection criteria (in order of priority):
    1. Date in filename (12 31 2024 = December 31, 2024)
    2. Major revision number (Rev156 > Rev153)
    3. Minor revision number (Rev156-3 > Rev156-1)
    """
    # Pattern: NLTS-PR FS MM DD YYYY RevXXX-Y.xlsm
    pattern = re.compile(
        r'NLTS-PR FS (\d{1,2}) (\d{1,2}) (\d{4}) Rev(\d+)-(\d+)\.xlsm$', 
        re.IGNORECASE
    )
    
    candidates = []
    for filename in os.listdir(NLTS_PR_DIR):
        match = pattern.match(filename)
        if match:
            month = int(match.group(1))
            day = int(match.group(2))
            year = int(match.group(3))
            rev_major = int(match.group(4))
            rev_minor = int(match.group(5))
            
            try:
                file_date = datetime(year, month, day)
                filepath = os.path.join(NLTS_PR_DIR, filename)
                candidates.append({
                    'date': file_date,
                    'rev_major': rev_major,
                    'rev_minor': rev_minor,
                    'filepath': filepath,
                    'filename': filename
                })
            except ValueError:
                continue
    
    if not candidates:
        raise FileNotFoundError("No NLTS-PR FS Excel files found in D:\\NLTS-PR")
    
    # Sort by: date DESC, rev_major DESC, rev_minor DESC
    candidates.sort(
        key=lambda x: (x['date'], x['rev_major'], x['rev_minor']), 
        reverse=True
    )
    
    latest = candidates[0]
    print(f"Found {len(candidates)} FS file(s):")
    for c in candidates[:5]:  # Show top 5
        print(f"  - {c['filename']} (Rev{c['rev_major']}-{c['rev_minor']})")
    print(f"Selected: {latest['filename']} (Rev{latest['rev_major']}-{latest['rev_minor']}, dated {latest['date'].strftime('%B %d, %Y')})")
    
    return latest['filepath'], latest['date']


def determine_status(ratio_name, value):
    """
    Determine if ratio is good, warning, or danger based on benchmarks.
    Also returns the industry benchmark for comparison.
    Returns: (status, industry_value, industry_label)
    """
    name_lower = ratio_name.lower()
    for key, bench in BENCHMARKS.items():
        if key in name_lower:
            industry_val = bench.get('industry', None)
            industry_label = bench.get('industry_label', 'Industry')
            
            # If no thresholds defined, return neutral status
            if bench['good'] is None or bench['warning'] is None:
                return 'neutral', industry_val, industry_label
            
            if bench['higher_is_better']:
                if value >= bench['good']:
                    status = 'good'
                elif value >= bench['warning']:
                    status = 'warning'
                else:
                    status = 'danger'
            else:
                if value <= bench['good']:
                    status = 'good'
                elif value <= bench['warning']:
                    status = 'warning'
                else:
                    status = 'danger'
            
            return status, industry_val, industry_label
    
    return 'neutral', None, None


def extract_ratios_from_excel(filepath, file_date):
    """Extract ratios from the Ratios sheet of the Excel file."""
    print(f"Opening workbook: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    if 'Ratios' not in wb.sheetnames:
        raise ValueError("No 'Ratios' sheet found in workbook")
    
    ws = wb['Ratios']
    
    structured_ratios = {
        'company': 'National Lift Truck Service of PR, Inc.',
        'asOf': file_date.strftime('%B %d, %Y'),
        'source': os.path.basename(filepath),
        'extractedAt': datetime.now().isoformat(),
        'solvencyRatios': [],
        'safetyRatios': [],
        'profitabilityRatios': [],
        'assetManagementRatios': [],
        'leadingIndicators': []
    }
    
    current_category = 'solvencyRatios'
    
    for row in ws.iter_rows(min_row=1, max_row=150, max_col=8, values_only=True):
        # Skip empty rows
        if not any(row):
            continue
        
        # Get values from columns
        col_a = row[0] if len(row) > 0 else None  # Index or category header
        col_b = row[1] if len(row) > 1 else None  # Ratio name
        col_c = row[2] if len(row) > 2 else None  # Formula or value
        col_d = row[3] if len(row) > 3 else None  # Current value
        col_e = row[4] if len(row) > 4 else None  # Prior value
        
        # Detect category headers ONLY when this is NOT a ratio row (col_a is text, not number)
        # Header rows have text in col_a like "SOLVENCY RATIOS", "PROFITABILITY RATIOS"
        # Ratio rows have an index number (1, 2, 3...) in col_a
        is_header_row = isinstance(col_a, str) or col_a is None or (isinstance(col_a, (int, float)) and col_a > 30)
        
        if is_header_row:
            for val in [col_a, col_b]:
                if isinstance(val, str):
                    upper = val.upper()
                    if 'SOLVENCY' in upper:
                        current_category = 'solvencyRatios'
                        break
                    elif 'SAFETY' in upper:
                        current_category = 'safetyRatios'
                        break
                    elif 'PROFITABILIT' in upper:  # Match "PROFITABILITY" specifically, not "Net Profit Margin"
                        current_category = 'profitabilityRatios'
                        break
                    elif 'ASSET MANAGEMENT' in upper or 'ASSET MGT' in upper:
                        current_category = 'assetManagementRatios'
                        break
                    elif 'LEADING' in upper or 'FINANCIAL INDICATOR' in upper:
                        current_category = 'leadingIndicators'
                        break
        
        # Check if this is a ratio row (has numeric index in col_a)
        # Ratios in the Excel have format: [index, name, formula_hint, current_value, prior_value, ...]
        is_ratio_row = isinstance(col_a, (int, float)) and col_a > 0 and col_a <= 30
        
        # Check for leading indicators - they may have name in col_a or col_b
        # These include: Break-even, Sustainable Growth, Z-Score, Retained Earnings, Working Capital, EBITDA, etc.
        is_leading_indicator = False
        leading_indicator_keywords = [
            'variable cost % of sales',
            'contribution margin',
            'break-even as % of net sales',
            '% break-even sales',
            'sustainable growth - current',
            'sustainable growth - no new debt',
            'z-score',
            'retained earnings to total assets',
            'working capital to total assets',
            'ebitda'
        ]
        
        if current_category == 'leadingIndicators':
            # Check col_b for indicator name (most common pattern: col_a is None, name in col_b)
            if isinstance(col_b, str) and len(col_b.strip()) > 3:
                check_str = col_b.lower().strip()
                
                # Check if this is a main indicator (not a sub-component)
                # Main indicators: have name without '=' AND have a value in col_d
                has_value = isinstance(col_d, (int, float)) and col_d != 0
                is_formula_row = '=' in col_b
                
                if has_value and not is_formula_row:
                    # Check against our keywords
                    for keyword in leading_indicator_keywords:
                        if keyword in check_str:
                            is_leading_indicator = True
                            break
                    
                    # Also catch indicators with parentheses like "Contribution Margin (%)"
                    if not is_leading_indicator:
                        # Check for patterns without exact match
                        if ('variable cost' in check_str and 'sales' in check_str) or \
                           ('contribution margin' in check_str) or \
                           ('break-even' in check_str or 'break even' in check_str) or \
                           ('sustainable growth' in check_str) or \
                           ('z-score' in check_str or 'bankruptcy' in check_str) or \
                           ('retained earnings' in check_str and 'total assets' in check_str) or \
                           ('working capital' in check_str and 'total assets' in check_str) or \
                           (check_str == 'ebitda'):
                            is_leading_indicator = True
        
        if not (is_ratio_row or is_leading_indicator):
            continue
        
        # Extract ratio name - check col_b first, then col_a for leading indicators
        ratio_name = None
        
        # For leading indicators, name might be in col_a
        if is_leading_indicator and isinstance(col_a, str) and len(col_a) > 3:
            check_lower = col_a.lower()
            # Check if it matches our keywords
            for keyword in leading_indicator_keywords:
                if keyword in check_lower:
                    ratio_name = col_a.strip()
                    break
        
        # If not found, check col_b
        if not ratio_name and isinstance(col_b, str) and len(col_b) > 3:
            # Clean up the name - remove formula parts
            if '=' in col_b:
                ratio_name = col_b.split('=')[0].strip()
            else:
                ratio_name = col_b.strip()
        
        # If still no name but we have col_a as text (for leading indicators)
        if not ratio_name and is_leading_indicator and isinstance(col_a, str) and len(col_a.strip()) > 3:
            if '=' not in col_a:
                ratio_name = col_a.strip()
        
        if not ratio_name:
            continue
            
        # Get the numeric value - check col_d first, then col_c (for leading indicators with different structure)
        current_val = None
        prior_val = None
        
        # Try col_d first
        if isinstance(col_d, (int, float)) and col_d != 0:
            current_val = col_d
        # If no value in col_d, try col_c for leading indicators
        elif is_leading_indicator and isinstance(col_c, (int, float)) and col_c != 0:
            current_val = col_c
        
        if isinstance(col_e, (int, float)):
            prior_val = col_e if col_e != 0 else None
        
        # Include if we have a valid ratio name and value
        if ratio_name and current_val is not None:
            current_rounded = round(current_val, 4)
            prior_rounded = round(prior_val, 4) if prior_val else None
            status, industry_val, industry_label = determine_status(ratio_name, current_rounded)
            
            entry = {
                'name': ratio_name,
                'current': current_rounded,
                'prior': prior_rounded if prior_rounded else 'N/A',
                'status': status,
                'industryBenchmark': industry_val,
                'industryLabel': industry_label
            }
            structured_ratios[current_category].append(entry)
    
    wb.close()
    
    # Remove duplicates by name within each category
    for cat in ['solvencyRatios', 'safetyRatios', 'profitabilityRatios', 'assetManagementRatios']:
        seen = set()
        unique_list = []
        for r in structured_ratios[cat]:
            name_key = r['name'].lower()
            if name_key not in seen:
                seen.add(name_key)
                unique_list.append(r)
        structured_ratios[cat] = unique_list
    
    # Special processing for leading indicators
    # 1. Filter out Z-Score subcomponents (start with '+' or '*')
    # 2. Remove duplicates (keep first/best match)
    # 3. Filter out rows that are clearly subcomponents
    leading = structured_ratios.get('leadingIndicators', [])
    filtered_leading = []
    leading_seen = set()
    
    for r in leading:
        name = r['name']
        name_lower = name.lower().strip()
        
        # Skip Z-Score subcomponents (like "+ Retained Earnings to Total Assets * 1.4")
        if name.strip().startswith('+') or name.strip().startswith('-'):
            continue
        if '*' in name and any(c.isdigit() for c in name.split('*')[-1]):
            continue
        
        # Normalize name for deduplication
        # e.g., "Contribution Margin (%)" and "Contribution Margin" are the same
        base_name = name_lower.replace('(%)', '').replace('(%', '').strip()
        
        if base_name not in leading_seen:
            leading_seen.add(base_name)
            filtered_leading.append(r)
    
    structured_ratios['leadingIndicators'] = filtered_leading
    
    return structured_ratios


def main():
    print("=== NLT-PR Dynamic Ratio Extractor ===\n")
    
    # Find latest file
    filepath, file_date = find_latest_fs_file()
    
    # Extract ratios
    ratios = extract_ratios_from_excel(filepath, file_date)
    
    # Count totals
    total_ratios = sum(len(ratios[cat]) for cat in ['solvencyRatios', 'safetyRatios', 'profitabilityRatios', 'assetManagementRatios'])
    
    # Print summary
    print(f"\n=== Extracted {total_ratios} Ratios (as of {ratios['asOf']}) ===")
    for cat in ['solvencyRatios', 'safetyRatios', 'profitabilityRatios', 'assetManagementRatios']:
        if ratios[cat]:
            print(f"\n{cat} ({len(ratios[cat])}):")
            for r in ratios[cat]:
                status_icon = '[OK]' if r['status'] == 'good' else ('[!]' if r['status'] == 'warning' else '[x]')
                print(f"  {status_icon} {r['name']}: {r['current']} (prior: {r['prior']})")
    
    # Save to JSON
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(ratios, f, indent=2, ensure_ascii=False)
    
    print(f"\n[OK] Saved to {OUTPUT_FILE}")
    return ratios


if __name__ == '__main__':
    main()
